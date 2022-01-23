# coding: utf-8

from pymongo import MongoClient
from openpyxl import load_workbook
import os, pprint, json


client = MongoClient()
db = client.standardDatabase

def MongoDB_Store_Write(filepath):
    '''
    :filepath: absolute path of base data(".xlsx" or ".txt")
    '''
    for root, dir, files in os.walk(filepath):
        for filename in files:
            if ".txt" in filename:
                with open(root + "/" + filename, "r") as f:
                    list_page_elements = []
                    page_elements = [e.strip("\n") for e in f.readline().split(",")]
                    for i in f.readlines():
                        list_page_elements.append([j.strip("\n") for j in i.split(",")])
                for row in list_page_elements:
                    dict_page_elements = {}
                    for k,v in zip(page_elements,row):
                        dict_page_elements[k]=float(v)
                    post = {"standard_name":filename.split(",")[0],
                        "pages":filename.split(",")[1],
                        "page_title": filename.split(",")[2].rstrip(".txt"),
                        "standard_data":dict_page_elements
                    }
                    db.standards.insert_one(post)
            elif ".xlsx" in filename:
                wb = load_workbook(root + "/" + filename)
                ws = wb.active
                list_page_elements = []
                for row in ws.iter_rows(min_row=1,max_row=1,values_only=True):
                    page_elements = row
                for else_row in ws.iter_rows(min_row=2,values_only=True):
                    list_page_elements.append(else_row)
                for row in list_page_elements:
                    dict_page_elements = {}
                    for k,v in zip(page_elements,row):
                        dict_page_elements[k]=float(v)
                    post = {"standard_name":filename.split(",")[0],
                        "pages":filename.split(",")[1],
                        "page_title": filename.split(",")[2].rstrip(".xlsx"),
                        "standard_data":dict_page_elements
                    }
                    db.standards.insert_one(post)
                    
def Json_Store_Write(filepath):
    '''
    :filepath: absolute path of base data(".xlsx" or ".txt")
    '''
    for root, dir, files in os.walk(filepath):
        for filename in files:
            if ".txt" in filename:
                list_page_elements = []
                with open(root + "/" + filename, "r") as f:
                    page_elements = [e.strip("\n") for e in f.readline().split(",")]
                    for i in f.readlines():
                        list_page_elements.append([j.strip("\n") for j in i.split(",")])
                for row_content in list_page_elements:
                    dict_page_elements = {}
                    for k,v in zip(page_elements,row_content):
                        dict_page_elements[k]=float(str(v).replace(" ",""))
                    post = {"standard_name":filename.split(",")[0],
                        "pages":filename.split(",")[1],
                        "page_title": filename.split(",")[2].rstrip(".txt"),
                        "standard_data":dict_page_elements
                    }
                    with open(root + "/" + "standarddata.json",'a+') as json_file:
                        json.dump(post, json_file)
            elif ".xlsx" in filename:
                wb = load_workbook(root + "/" + filename)
                ws = wb.active
                list_page_elements = []
                for row in ws.iter_rows(min_row=1,max_row=1,values_only=True):
                    page_elements = row
                for else_row in ws.iter_rows(min_row=2,values_only=True):
                    list_page_elements.append(else_row)
                for row in list_page_elements:
                    dict_page_elements = {}
                    for k,v in zip(page_elements,row):
                        dict_page_elements[k]=float(str(v).replace(" ",""))
                    post = {"standard_name":filename.split(",")[0],
                        "pages":filename.split(",")[1],
                        "page_title": filename.split(",")[2].rstrip(".xlsx"),
                        "standard_data":dict_page_elements
                    }
                    with open(root + "/" + "standarddata.json",'a+') as json_file:
                        json.dump(post, json_file)


if __name__ == "__main__":
    filepath = input("请输入原始数据文件绝对路径: ")
    Json_Store_Write(filepath=filepath)